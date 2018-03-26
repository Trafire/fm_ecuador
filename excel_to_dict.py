from openpyxl import load_workbook

def get_dict(filename):
    wb = load_workbook(filename = filename)
    ws = wb["standing"]
    index = 0
    headings = []
    products = []
    for row in ws.rows:
        col = 0
        line = {}
        for cell in row:
            if cell.value == None:
                break
            if index == 0:
                headings.append(cell.value)
            else:
                line[headings[col]] = cell.value
            col += 1
        if index != 0:
            products.append(line)
        index += 1
    return products

def get_dict_gen(filename):
    wb = load_workbook(filename = filename)
    ws = wb.active
    index = 0
    headings = []
    products = []
    for row in ws.rows:
        col = 0
        line = {}
        for cell in row:
            if cell.value == None:
                break
            if index == 0:
                headings.append(cell.value)
            else:
                line[headings[col]] = cell.value
            col += 1
        if index != 0:
            products.append(line)
        index += 1
    return products
